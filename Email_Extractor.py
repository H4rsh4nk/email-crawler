from File_Write import File_Write
from File_Read import File_Read
from Core import Core
from openpyxl import load_workbook





if __name__ == "__main__":
    _File_Write = File_Write()
    _File_Read = File_Read()
    _Core = Core()

    # print(" ___                       ___     ___  __        __  ___  __   __")
    # print("|__   |\/|  /\  | |    __ |__  \_/  |  |__)  /\  /  `  |  /  \ |__)")
    # print("|___  |  | /~~\ | |___    |___ / \  |  |  \ /~~\ \__,  |  \__/ |  \ \n")
    # from google.colab import files
    # excel_file = files.upload()
    # excel_file = "Book.xlsx"
    # try:
    excel_file = input("Enter file name : " + ".xlsx") 

    wb = load_workbook(excel_file)
    sh1 = wb['Sheet1']
    row = sh1.max_row
    column = sh1.max_column
    tmp = sh1.cell(1,5).value

    for i in range(1 , row+1):
        sh1.cell(i,6).value = sh1.cell(i,6).value.replace('Registered Member ', '')

        if(sh1.cell(i,4).value == ""):
            continue
        # else:
        elif((sh1.cell(i,5).value == "") or (sh1.cell(i,5).value == None)): #or (sh1.cell(i,5).value == ""))):
            _Core.URL(sh1.cell(i,4).value)
            # print(_Core._Core__Emails)
            if( _Core._Core__Emails != []):
                sh1.cell(i,5).value = _Core._Core__Emails
                pe = (i/row)*100
                print(str("{:.2f}".format(pe)) + "% " + "Entry : " + str(i) + "  |" + _Core._Core__Emails)
        wb.save(excel_file)   
    print("All entries are done")
    for i in range(2 , row+1):
        if(sh1.cell(i,5).value == tmp):
            sh1.cell(i,5).value = ""
        else:
            tmp = sh1.cell(i,5).value
    wb.save("Book.xlsx") 
    # Write text to file
    # _File_Write.File_W(set(_Core._Core__Emails))
# except Exception:
#     print("error")

#     excel_file = "C:\\Users\\patha\\Downloads\\Therapistsspecialisi.xlsx"
#     # try:
#     wb = load_workbook(excel_file)
#     sh1 = wb['Sheet1']
#     row = sh1.max_row
#     column = sh1.max_column
#     tmp = sh1.cell(1,5).value

#     for i in range(1 , row+1):
#         sh1.cell(i,6).value = sh1.cell(i,6).value.replace('Registered Member ', '')

#         if(sh1.cell(i,4).value == ""):
#             continue
#         # else:
#         elif((sh1.cell(i,5).value == "") or (sh1.cell(i,5).value == None)): #or (sh1.cell(i,5).value == ""))):
#             _Core.URL(sh1.cell(i,4).value)
#             # print(_Core._Core__Emails)
#             if( _Core._Core__Emails != []):
#                 sh1.cell(i,5).value = _Core._Core__Emails
#                 pe = (i/row)*100
#                 print(str("{:.2f}".format(pe)) + "% " + "Entry : " + str(i) + "  |" + _Core._Core__Emails)
#         wb.save(excel_file)   
#     print("All entries are done")
#     for i in range(2 , row+1):
#         if(sh1.cell(i,5).value == tmp):
#             sh1.cell(i,5).value = ""
#         else:
#             tmp = sh1.cell(i,5).value
#     wb.save("Book.xlsx") 
#     # from google.colab import files
#     # files.download(excel_file)
#     # Write text to file
#     # _File_Write.File_W(set(_Core._Core__Emails))
# # except Exception:
# #     print("error")
