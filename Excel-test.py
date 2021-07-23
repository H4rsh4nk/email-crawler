from openpyxl import load_workbook

wb = load_workbook("Book2.xlsx")
sh1 = wb['Sheet1']
row = sh1.max_row
column = sh1.max_column
tmp = sh1.cell(1,5).value
for i in range(2 , row+1):
    if(sh1.cell(i,5).value == tmp):
        sh1.cell(i,5).value = ""
    else:
        tmp = sh1.cell(i,5).value
wb.save("Book.xlsx")
        
    
#     if(sh1.cell(i,1).value == "Harshank"):
#         sh1.cell(i,2).value = "1"

# -------------------------------------------------------------------------------------------
# from openpyxl import load_workbook

# wb = load_workbook("Book2.xlsx")
# ws = wb.active

# for row in ws.rows:
#     if row[0].value == "Harshank":
#         print(row)
#         print(row[1])

#         for cell in row:
#             print(cell.value, end=" ")
#         print()
# -------------------------------------------------------------------------------------------
# import openpyxl
# xfile = openpyxl.load_workbook('Book2.xlsx')

# sheet = xfile.get_sheet_by_name('Sheet1')
# sheet['A2'] = 'hello world'
# xfile.save('Book2.xlsx')

# -------------------------------------------------------------------------------------------
# import xlsxwriter

# # Create a workbook and add a worksheet.
# workbook = xlsxwriter.Workbook('Book2.xlsx')
# worksheet = workbook.add_worksheet()

# # Some data we want to write to the worksheet.
# expenses = (
#     ['Harshank', 1000],
#     ['Gas',   100],
#     ['Food',  300],
#     ['Gym',    50],
# )

# # Start from the first cell. Rows and columns are zero indexed.
# row = 4
# col = 0

# # Iterate over the data and write it out row by row.
# for item, cost in (expenses):
#     worksheet.write(row, col,     item)
#     worksheet.write(row, col + 1, cost)
#     row += 1

# # Write a total using a formula.
# worksheet.write(row, 0, 'Total')
# worksheet.write(row, 1, '=SUM(B1:B4)')

# workbook.close()